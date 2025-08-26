import streamlit as st
import re
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from io import BytesIO
import os

# Ustawienie layoutu strony
st.set_page_config(layout="wide")

# Tytuł aplikacji
st.title("🗺️ Generator pliku excel - MAPA PRZEDMIOTÓW")
st.write("Wgraj plik excel z twoją bazą danych według ustalonych kryteriów.")

# Sekcja do przesyłania pliku
st.markdown("### 1. Prześlij plik Excel (.xlsx)")
uploaded_file = st.file_uploader("Wybierz plik", type=["xlsx"])

if uploaded_file:
    try:
        # Zapisanie pliku tymczasowo na dysku
        temp_file_path = "temp_uploaded_file.xlsx"
        with open(temp_file_path, "wb") as f:
            f.write(uploaded_file.getbuffer())

        # Funkcja przetwarzająca plik z dysku
        def process_excel(filepath):
            try:
                wb = load_workbook(filepath)
                ws_baza = wb["Baza"]
                ws_stara = wb["STARA mapa"]
            except KeyError as e:
                st.error(f"Błąd: W pliku brakuje wymaganego arkusza: {e}. Upewnij się, że plik zawiera arkusze 'Baza' i 'STARA mapa'.")
                return None, None, None

            # Funkcje pomocnicze
            def clean_nazwisko(nazwisko):
                if not nazwisko:
                    return ""
                return re.sub(r"\(.*?\)", "", str(nazwisko)).strip().lower()

            # Mapa maili
            mail_map = {}
            for row in ws_stara.iter_rows(min_row=2, values_only=True):
                imie = str(row[2]).strip() if row[2] else ""
                nazwisko = str(row[3]).strip() if row[3] else ""
                mail = str(row[1]).strip() if row[1] else ""
                if imie and nazwisko and mail:
                    key = f"{imie.lower()} {nazwisko.lower()}"
                    mail_map[key] = mail

            def find_mail_safe(imie, nazwisko, mail_map):
                key_full = f"{imie.lower()} {nazwisko.lower()}"
                if key_full in mail_map:
                    return mail_map[key_full]
                candidates = []
                for k, mail in mail_map.items():
                    k_imie, k_nazwisko = k.split(" ", 1)
                    k_nazwisko_clean = clean_nazwisko(k_nazwisko)
                    if imie.lower() == k_imie and nazwisko.lower() == k_nazwisko_clean:
                        candidates.append(mail)
                if len(candidates) == 1:
                    return candidates[0]
                else:
                    return "BRAK MAILA"

            # Nowy skoroszyt i arkusze w pamięci
            new_wb = Workbook()
            ws_mapa = new_wb.active
            ws_mapa.title = "MAPA PRZEDMIOTÓW"
            ws_brak_id = new_wb.create_sheet("BRAKUJĄCE ID")
            ws_brak_mail = new_wb.create_sheet("BRAKUJĄCE EMAILE")

            # Kolory i obramowania
            color_map = {"LO": "ADD8E6", "1-3 SP": "FFFF99", "4-8 SP": "CCFFCC"}
            green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            red_fill = PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type="solid")
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))

            # Nagłówki MAPA PRZEDMIOTÓW
            headers = ["ID", "Email", "Imię", "Nazwisko", "Przedmiot", "Poziom edukacyjny", "Rozszerzenie", "Aktywny"]
            for col_num, header in enumerate(headers, 1):
                cell = ws_mapa.cell(row=1, column=col_num, value=header)
                cell.border = thin_border
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Arkusz BRAKUJĄCE ID
            ws_brak_id.merge_cells('A1:B1')
            cell = ws_brak_id['A1']
            cell.value = "Osoby, które nie mają ID"
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws_brak_id.append(["Imię", "Nazwisko"])
            for c in ws_brak_id[2]:
                c.font = Font(bold=True)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = thin_border

            # Arkusz BRAKUJĄCE EMAILE
            ws_brak_mail.merge_cells('A1:B1')
            cell = ws_brak_mail['A1']
            cell.value = "Osoby, które nie mają Email"
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws_brak_mail.append(["Imię", "Nazwisko"])
            for c in ws_brak_mail[2]:
                c.font = Font(bold=True)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = thin_border

            brak_id_set = set()
            brak_mail_set = set()
            row_out = 2

            for row in ws_baza.iter_rows(min_row=2, values_only=True):
                _id = row[0]
                imie = str(row[1]).strip() if row[1] else ""
                nazwisko = str(row[2]).strip() if row[2] else ""
                sp_subjects = row[4]
                lo_subjects = row[5]

                mail = find_mail_safe(imie, nazwisko, mail_map)

                # Przetwarzanie przedmiotów SP
                if sp_subjects:
                    for subj in [s.strip() for s in str(sp_subjects).split(",")]:
                        if subj:
                            typ = "1-3 SP" if subj.lower() == "edukacja wczesnoszkolna" else "4-8 SP"
                            values = [_id, mail, imie, nazwisko, subj, typ, "NIE", "TAK"]
                            
                            if not _id:
                                brak_id_set.add((imie, nazwisko))
                            if mail == "BRAK MAILA":
                                brak_mail_set.add((imie, nazwisko))

                            for col_num, val in enumerate(values, 1):
                                cell = ws_mapa.cell(row=row_out, column=col_num, value=val)
                                if col_num == 6 and typ in color_map:
                                    cell.fill = PatternFill(start_color=color_map[typ], end_color=color_map[typ], fill_type="solid")
                                if col_num == 7:
                                    cell.fill = red_fill
                                if col_num == 8:
                                    cell.fill = green_fill
                                cell.border = thin_border
                            row_out += 1

                # Przetwarzanie przedmiotów LO
                if lo_subjects:
                    for subj in [s.strip() for s in str(lo_subjects).split(",")]:
                        if subj:
                            typ = "LO"
                            values = [_id, mail, imie, nazwisko, subj, typ, "TAK", "TAK"]

                            if not _id:
                                brak_id_set.add((imie, nazwisko))
                            if mail == "BRAK MAILA":
                                brak_mail_set.add((imie, nazwisko))

                            for col_num, val in enumerate(values, 1):
                                cell = ws_mapa.cell(row=row_out, column=col_num, value=val)
                                if col_num == 6 and typ in color_map:
                                    cell.fill = PatternFill(start_color=color_map[typ], end_color=color_map[typ], fill_type="solid")
                                if col_num == 7:
                                    cell.fill = green_fill
                                if col_num == 8:
                                    cell.fill = green_fill
                                cell.border = thin_border
                            row_out += 1

            # Dodanie brakujących do arkuszy z obramowaniami
            for imie, nazwisko in brak_id_set:
                ws_brak_id.append([imie, nazwisko])
                for c in ws_brak_id[ws_brak_id.max_row]:
                    c.border = thin_border
            for imie, nazwisko in brak_mail_set:
                ws_brak_mail.append([imie, nazwisko])
                for c in ws_brak_mail[ws_brak_mail.max_row]:
                    c.border = thin_border

            # Zapisanie skoroszytu do bufora
            output_buffer = BytesIO()
            new_wb.save(output_buffer)
            output_buffer.seek(0)
            return output_buffer, list(brak_id_set), list(brak_mail_set)

        # Uruchomienie przetwarzania i wyświetlenie wyników
        with st.spinner('Przetwarzam plik...'):
            output_buffer, brak_id, brak_mail = process_excel(temp_file_path)

        # Usunięcie pliku tymczasowego
        os.remove(temp_file_path)

        if output_buffer:
            st.success("Plik został pomyślnie przetworzony! 🎉")

            st.markdown("### 2. Pobierz gotowy plik")
            st.download_button(
                label="Pobierz plik wynikowy",
                data=output_buffer,
                file_name="wynik.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            # Wyświetlenie informacji o brakach
            st.markdown("---")
            st.markdown("### 3. Podsumowanie")
            
            if brak_id:
                st.warning(f"Znaleziono {len(brak_id)} osób bez ID w pliku 'Baza'.")
                df_brak_id = pd.DataFrame(brak_id, columns=["Imię", "Nazwisko"])
                st.markdown("Oto lista osób, które nie mają ID:")
                st.dataframe(df_brak_id, use_container_width=True)

            if brak_mail:
                st.warning(f"Znaleziono {len(brak_mail)} osób bez maila w pliku 'STARA mapa'.")
                df_brak_mail = pd.DataFrame(brak_mail, columns=["Imię", "Nazwisko"])
                st.markdown("Oto lista osób, które nie mają emaila:")
                st.dataframe(df_brak_mail, use_container_width=True)

    except Exception as e:
        st.error(f"Wystąpił błąd podczas przetwarzania pliku: {e}")
